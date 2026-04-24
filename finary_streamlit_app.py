import io
import re
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st


st.set_page_config(page_title="Finary Care Ops Analyzer", layout="wide")

DEFAULT_RULES = [
    {
        "name": "Account sync reliability",
        "owner": "Engineering + Product",
        "type": "Product reliability",
        "patterns": [
            r"bank[\s_-]?sync",
            r"aggregation",
            r"aggregator",
            r"connect(ion|ed|ing)?\s+bank",
            r"account\s+disconnect",
            r"missing\s+transaction",
            r"balance\s+(wrong|mismatch|incorrect|stale)",
            r"reconnect",
            r"sync\s+fail",
            r"sync\s+issue",
            r"account\s+sync",
            r"bank\s+connection",
        ],
    },
    {
        "name": "Billing and subscription clarity",
        "owner": "Product + Ops",
        "type": "Product + Process",
        "patterns": [
            r"billing",
            r"subscription",
            r"charged",
            r"charge",
            r"refund",
            r"cancel",
            r"cancellation",
            r"invoice",
            r"payment\s+method",
            r"plan",
            r"renewal",
            r"trial",
            r"pricing",
        ],
    },
    {
        "name": "Finary One service delivery and SLA",
        "owner": "Ops + Product",
        "type": "Service operations",
        "patterns": [
            r"finary\s+one",
            r"advisor",
            r"wealth",
            r"premium",
            r"appointment",
            r"meeting",
            r"call\s+back",
            r"service\s+delay",
            r"sla",
            r"onboarding",
            r"portfolio\s+review",
            r"private\s+client",
        ],
    },
    {
        "name": "Product bug stability",
        "owner": "Engineering",
        "type": "Product stability",
        "patterns": [
            r"bug",
            r"crash",
            r"error",
            r"not\s+working",
            r"doesn'?t\s+work",
            r"broken",
            r"freeze",
            r"loading",
            r"display\s+issue",
            r"wrong\s+value",
            r"duplicate",
            r"issue\s+in\s+app",
            r"problem\s+in\s+app",
        ],
    },
    {
        "name": "Tax and reporting support complexity",
        "owner": "Ops + Product",
        "type": "Support complexity",
        "patterns": [
            r"tax",
            r"fiscal",
            r"report",
            r"pea",
            r"ifi",
            r"declaration",
            r"csv",
            r"export",
            r"statement",
            r"year[-\s]?end",
            r"document",
        ],
    },
    {
        "name": "Security and account access friction",
        "owner": "Security + Ops",
        "type": "Trust and access",
        "patterns": [
            r"password",
            r"login",
            r"sign[\s-]?in",
            r"2fa",
            r"two[\s-]?factor",
            r"otp",
            r"suspicious",
            r"hacked",
            r"fraud",
            r"access",
            r"locked",
            r"verification",
        ],
    },
]

VIRALITY_PATTERNS_DEFAULT = [
    r"scam",
    r"unacceptable",
    r"ridiculous",
    r"angry",
    r"furious",
    r"terrible",
    r"awful",
    r"complain",
    r"twitter",
    r"linkedin",
    r"social",
    r"public",
    r"frustrat",
    r"annoy",
    r"disappointed",
    r"trust",
    r"urgent",
    r"immediately",
]

TRUST_RISK_PATTERNS_DEFAULT = [
    r"money",
    r"balance",
    r"net worth",
    r"portfolio",
    r"charged",
    r"refund",
    r"missing transaction",
    r"bank",
    r"security",
    r"access",
    r"hacked",
]

PREMIUM_TIERS_DEFAULT = ["plus", "premium", "gold", "finary one", "one", "pro", "private"]
HIGH_PRIORITY_VALUES_DEFAULT = ["high", "urgent", "p1", "critical"]
OPEN_STATUS_VALUES_DEFAULT = ["open", "pending", "waiting", "in progress", "new"]

EXPECTED_COLS = [
    "ticket_id",
    "created_at",
    "channel",
    "customer_id",
    "customer_email",
    "country",
    "subscription_tier",
    "language",
    "category",
    "subcategory",
    "priority",
    "subject",
    "description",
    "status",
    "first_response_time_min",
    "resolution_time_hours",
    "csat",
    "escalated_to_product",
    "tags",
]


# -------------------------
# Helpers
# -------------------------
def normalize_text(x):
    if x is None:
        return ""
    if isinstance(x, float) and np.isnan(x):
        return ""
    return str(x).strip().lower()


def safe_float(x, default=0.0):
    try:
        if x is None or str(x).strip() == "":
            return default
        return float(x)
    except Exception:
        return default


def min_max_score(series: pd.Series, scale_max: float = 10.0) -> pd.Series:
    if len(series) == 0:
        return pd.Series(dtype=float)
    if series.nunique(dropna=False) <= 1:
        return pd.Series([scale_max / 2] * len(series), index=series.index)
    min_v = series.min()
    max_v = series.max()
    if max_v == min_v:
        return pd.Series([scale_max / 2] * len(series), index=series.index)
    return ((series - min_v) / (max_v - min_v) * scale_max).round(2)


def parse_multiline_list(text: str) -> List[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def parse_tags(text: str) -> List[str]:
    raw = normalize_text(text)
    if not raw:
        return []
    parts = re.split(r"[,;|/\\]+", raw)
    cleaned = []
    for p in parts:
        token = p.strip()
        if token:
            cleaned.append(token)
    return cleaned


def tokens_from_evidence(text: str) -> List[str]:
    raw = normalize_text(text)
    if not raw:
        return []
    parts = re.split(r"[^a-zA-Z0-9_\-]+", raw)
    return [p for p in parts if len(p) >= 3]


@st.cache_data(show_spinner=False)
def read_excel_safe(file_bytes: bytes, filename: str) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes)
    try:
        return pd.read_excel(bio, sheet_name=0)
    except Exception:
        from openpyxl import load_workbook

        bio.seek(0)
        wb = load_workbook(bio, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = list(ws.values)
        if not data:
            return pd.DataFrame()
        headers = data[0]
        rows = data[1:]
        return pd.DataFrame(rows, columns=headers)


# -------------------------
# Rule system
# -------------------------
def build_evidence_text(row: pd.Series, include_tags: bool = True) -> str:
    parts = [
        normalize_text(row.get("category")),
        normalize_text(row.get("subcategory")),
        normalize_text(row.get("subject")),
        normalize_text(row.get("description")),
    ]
    if include_tags:
        parts.append(normalize_text(row.get("tags")))
    parts.append(normalize_text(row.get("subscription_tier")))
    return " | ".join([p for p in parts if p])


def classify_opportunity(evidence_text: str, rules: List[Dict]) -> Dict:
    matched = []
    for rule in rules:
        hits = [p for p in rule["patterns"] if re.search(p, evidence_text, flags=re.IGNORECASE)]
        if hits:
            matched.append(
                {
                    "name": rule["name"],
                    "owner": rule["owner"],
                    "type": rule["type"],
                    "match_count": len(hits),
                    "matched_patterns": hits,
                }
            )

    if not matched:
        return {
            "opportunity_name": "Other / uncategorized",
            "owner": "Ops",
            "opportunity_type": "Misc",
            "rule_strength": 0,
            "matched_patterns": "",
            "all_candidate_opportunities": "",
        }

    matched = sorted(matched, key=lambda x: x["match_count"], reverse=True)
    best = matched[0]
    candidates = [f"{m['name']} ({m['match_count']})" for m in matched]
    return {
        "opportunity_name": best["name"],
        "owner": best["owner"],
        "opportunity_type": best["type"],
        "rule_strength": best["match_count"],
        "matched_patterns": " | ".join(best["matched_patterns"]),
        "all_candidate_opportunities": " | ".join(candidates),
    }


def premium_flag(subscription_tier: str, premium_tiers: List[str]) -> int:
    return int(normalize_text(subscription_tier) in set([normalize_text(x) for x in premium_tiers]))


def high_priority_flag(priority: str, high_priority_values: List[str]) -> int:
    return int(normalize_text(priority) in set([normalize_text(x) for x in high_priority_values]))


def open_status_flag(status: str, open_status_values: List[str]) -> int:
    return int(normalize_text(status) in set([normalize_text(x) for x in open_status_values]))


def escalation_flag(x) -> int:
    if isinstance(x, bool):
        return int(x)
    return int(normalize_text(x) in {"true", "yes", "1", "y"})


def regex_hit_score(text: str, patterns: List[str], cap: int = 5) -> int:
    hits = sum(1 for p in patterns if re.search(p, text, flags=re.IGNORECASE))
    return min(hits, cap)


def severity_proxy(row: pd.Series) -> float:
    score = 0.0
    score += 2.0 * row.get("is_high_priority", 0)
    score += 1.5 * row.get("is_open", 0)
    score += 1.5 * row.get("is_escalated_product", 0)

    csat = row.get("csat_num", np.nan)
    if not np.isnan(csat):
        if csat <= 2:
            score += 2.0
        elif csat == 3:
            score += 1.0

    rt = row.get("resolution_time_hours_num", 0)
    if rt >= 72:
        score += 2.0
    elif rt >= 24:
        score += 1.0

    frt = row.get("first_response_time_min_num", 0)
    if frt >= 1440:
        score += 2.0
    elif frt >= 240:
        score += 1.0

    return round(score, 2)


def aggregate_opportunities(df: pd.DataFrame, weights: Dict[str, float]) -> pd.DataFrame:
    grouped = (
        df.groupby(["opportunity_name", "owner", "opportunity_type"], dropna=False)
        .agg(
            ticket_volume=("ticket_id", "count"),
            avg_rule_strength=("rule_strength", "mean"),
            avg_severity_proxy=("severity_proxy", "mean"),
            premium_share=("is_premium", "mean"),
            avg_virality_risk=("virality_risk", "mean"),
            avg_trust_risk=("trust_risk", "mean"),
            escalation_rate=("is_escalated_product", "mean"),
            avg_resolution_time_h=("resolution_time_hours_num", "mean"),
            avg_first_response_min=("first_response_time_min_num", "mean"),
            open_rate=("is_open", "mean"),
            tag_coverage=("has_tags", "mean"),
        )
        .reset_index()
    )

    if grouped.empty:
        return grouped

    grouped["volume_score"] = min_max_score(grouped["ticket_volume"], 10)
    grouped["severity_score"] = min_max_score(grouped["avg_severity_proxy"], 10)
    grouped["premium_score"] = (grouped["premium_share"] * 10).round(2)
    grouped["virality_score"] = min_max_score(grouped["avg_virality_risk"], 10)
    grouped["trust_score"] = min_max_score(grouped["avg_trust_risk"], 10)
    grouped["product_dependency_score"] = (grouped["escalation_rate"] * 10).round(2)
    grouped["ops_burden_score"] = min_max_score(grouped["avg_resolution_time_h"].fillna(0), 10)
    grouped["backlog_score"] = (grouped["open_rate"] * 10).round(2)

    weight_sum = sum(weights.values()) or 1.0
    grouped["opportunity_score"] = (
        grouped["volume_score"] * weights["volume"]
        + grouped["severity_score"] * weights["severity"]
        + grouped["premium_score"] * weights["premium"]
        + grouped["virality_score"] * weights["virality"]
        + grouped["product_dependency_score"] * weights["product_dependency"]
        + grouped["ops_burden_score"] * weights["ops_burden"]
        + grouped["backlog_score"] * weights["backlog"]
    ) / weight_sum
    grouped["opportunity_score"] = grouped["opportunity_score"].round(2)

    grouped = grouped.sort_values(["opportunity_score", "ticket_volume"], ascending=[False, False]).reset_index(drop=True)
    return grouped


def get_examples(df: pd.DataFrame, opportunity_name: str, n: int = 3) -> List[Dict]:
    subset = df[df["opportunity_name"] == opportunity_name].copy()
    if subset.empty:
        return []
    subset = subset.sort_values(
        by=["severity_proxy", "virality_risk", "resolution_time_hours_num"],
        ascending=[False, False, False],
    )
    examples = []
    for _, row in subset.head(n).iterrows():
        examples.append(
            {
                "ticket_id": row.get("ticket_id"),
                "subject": str(row.get("subject", "")),
                "category": str(row.get("category", "")),
                "subcategory": str(row.get("subcategory", "")),
                "subscription_tier": str(row.get("subscription_tier", "")),
                "tags": str(row.get("tags", "")),
            }
        )
    return examples


def top_tags_for_opportunity(df: pd.DataFrame, opportunity_name: str, top_n: int = 8) -> List[Tuple[str, int]]:
    subset = df[df["opportunity_name"] == opportunity_name].copy()
    tag_counts: Dict[str, int] = {}
    for text in subset["tags"].fillna("").astype(str).tolist():
        for tag in parse_tags(text):
            tag_counts[tag] = tag_counts.get(tag, 0) + 1
    return sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:top_n]


def default_report_text(opportunity_name: str) -> Tuple[str, str, List[str]]:
    if opportunity_name == "Account sync reliability":
        return (
            "Users lose trust when balances, connections, or transactions look wrong or stale.",
            "Core product credibility is at risk. High repeatability also creates avoidable support load.",
            [
                "Build connector-level reliability monitoring and stale-sync alerts.",
                "Improve reconnect flow and degraded-state messaging in-product.",
                "Create proactive user communication when a sync issue is known.",
                "Track top failing institutions and resolution speed weekly.",
            ],
        )
    if opportunity_name == "Finary One service delivery and SLA":
        return (
            "High-value customers expect concierge-level reliability and clear follow-up.",
            "Even lower volume can have outsized churn, reputation, and revenue impact because these users are premium.",
            [
                "Define explicit SLA by request type and make it operationally visible.",
                "Create premium queue routing and ownership rules.",
                "Add proactive updates for delayed advisor or onboarding requests.",
                "Review weekly premium pain points with Ops and Product.",
            ],
        )
    if opportunity_name == "Billing and subscription clarity":
        return (
            "Users feel charged unexpectedly or do not understand what plan actions do.",
            "This creates avoidable tickets, refund pressure, and conversion friction.",
            [
                "Clarify billing events and next charges in-app.",
                "Simplify cancel / refund / renewal explanations.",
                "Add self-serve billing FAQ and contextual help surfaces.",
                "Review the highest-friction subscription journeys with Product.",
            ],
        )
    if opportunity_name == "Product bug stability":
        return (
            "Bugs break trust and create repeated support demand when basic actions fail.",
            "Support becomes a workaround for defects that should be fixed upstream.",
            [
                "Create defect taxonomy by customer impact, not just technical cause.",
                "Prioritize bugs that affect trust, money, or repeated contacts.",
                "Feed recurring bug clusters into the product backlog weekly.",
                "Track ticket recurrence after release to validate fixes.",
            ],
        )
    return (
        "This theme creates repeated friction for users.",
        "It creates support load and likely reflects a product or process gap.",
        [
            "Document the recurring root causes.",
            "Create clear ownership with Product, Engineering, or Ops.",
            "Track recurrence after fixes are implemented.",
        ],
    )


def create_opportunity_reports(grouped: pd.DataFrame, df: pd.DataFrame) -> List[Dict]:
    reports = []
    top3 = grouped.head(3).copy()
    for _, row in top3.iterrows():
        opp = row["opportunity_name"]
        examples = get_examples(df, opp, n=3)
        top_tags = top_tags_for_opportunity(df, opp, top_n=6)
        user_impact, business_impact, asks = default_report_text(opp)
        justification = (
            f"This opportunity ranks highly because it combines repeatability (volume={int(row['ticket_volume'])}), "
            f"severity (avg severity proxy={row['avg_severity_proxy']:.2f}), revenue exposure "
            f"(premium share={row['premium_share']:.1%}), escalation rate ({row['escalation_rate']:.1%}), "
            f"open backlog ({row['open_rate']:.1%}), and operational burden "
            f"(avg resolution time={row['avg_resolution_time_h']:.1f}h)."
        )
        communication_approach = (
            "I would surface this to Product and Engineering as a recurring pattern, not as isolated anecdotes: "
            "quantify volume, customer trust impact, premium-user exposure, escalation rate, tag clusters, and examples "
            "of repeated failure modes. Then I would pair that evidence with a clear ask, owner, and success metric."
        )
        reports.append(
            {
                "opportunity_name": opp,
                "owner": row["owner"],
                "score": row["opportunity_score"],
                "ticket_volume": int(row["ticket_volume"]),
                "premium_share": float(row["premium_share"]),
                "escalation_rate": float(row["escalation_rate"]),
                "avg_severity_proxy": float(row["avg_severity_proxy"]),
                "avg_resolution_time_h": float(row["avg_resolution_time_h"]),
                "justification": justification,
                "user_impact": user_impact,
                "business_impact": business_impact,
                "communication_approach": communication_approach,
                "asks_to_product_engineering": asks,
                "examples": examples,
                "top_tags": top_tags,
            }
        )
    return reports


def diagnostics(df: pd.DataFrame) -> Dict:
    total = len(df)
    processed = int(df["opportunity_name"].notna().sum())
    uncategorized = int((df["opportunity_name"] == "Other / uncategorized").sum())
    tags_present = int(df["has_tags"].sum())
    matched_rules = int((df["rule_strength"] > 0).sum())
    duplicate_ticket_ids = int(df["ticket_id"].astype(str).duplicated().sum()) if "ticket_id" in df.columns else 0
    missing_subject = int(df["subject"].isna().sum())
    missing_description = int(df["description"].isna().sum())
    return {
        "total_tickets": total,
        "processed_tickets": processed,
        "coverage_rate": processed / total if total else 0.0,
        "uncategorized_tickets": uncategorized,
        "uncategorized_rate": uncategorized / total if total else 0.0,
        "tickets_with_tags": tags_present,
        "tag_coverage_rate": tags_present / total if total else 0.0,
        "matched_rules": matched_rules,
        "matched_rule_rate": matched_rules / total if total else 0.0,
        "duplicate_ticket_ids": duplicate_ticket_ids,
        "missing_subject": missing_subject,
        "missing_description": missing_description,
    }


def all_tag_counts(df: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    counts: Dict[str, int] = {}
    for text in df["tags"].fillna("").astype(str).tolist():
        for tag in parse_tags(text):
            counts[tag] = counts.get(tag, 0) + 1
    out = pd.DataFrame(sorted(counts.items(), key=lambda x: x[1], reverse=True), columns=["tag", "count"])
    return out.head(top_n)


def render_report_text(report: Dict) -> str:
    lines = [
        f"Opportunity: {report['opportunity_name']}",
        "",
        "Evidence:",
        f"- {report['ticket_volume']} tickets / total dataset",
        f"- Avg severity proxy: {report['avg_severity_proxy']:.2f}",
        f"- Premium share: {report['premium_share']:.1%}",
        f"- Escalation rate to Product: {report['escalation_rate']:.1%}",
        f"- Avg resolution time: {report['avg_resolution_time_h']:.1f}h",
    ]
    if report["top_tags"]:
        lines.append("- Top tags: " + ", ".join([f"{tag} ({count})" for tag, count in report["top_tags"]]))
    lines += [
        "",
        "User impact:",
        report["user_impact"],
        "",
        "Business impact:",
        report["business_impact"],
        "",
        "Ask to Product/Eng:",
    ]
    for i, ask in enumerate(report["asks_to_product_engineering"], start=1):
        lines.append(f"{i}. {ask}")
    return "\n".join(lines)


def build_markdown_reports(reports: List[Dict]) -> str:
    lines = ["# Top 3 Opportunity Reports", ""]
    for i, report in enumerate(reports, start=1):
        lines.append(f"## {i}. {report['opportunity_name']}")
        lines.append(f"- **Owner:** {report['owner']}")
        lines.append(f"- **Score:** {report['score']}")
        lines.append(f"- **Why prioritize:** {report['justification']}")
        lines.append(f"- **User impact:** {report['user_impact']}")
        lines.append(f"- **Business impact:** {report['business_impact']}")
        lines.append(f"- **Communication approach:** {report['communication_approach']}")
        if report["top_tags"]:
            lines.append("- **Top tags:** " + ", ".join([f"{tag} ({count})" for tag, count in report["top_tags"]]))
        lines.append("- **Recommended actions:**")
        for ask in report["asks_to_product_engineering"]:
            lines.append(f"  - {ask}")
        lines.append("- **Example tickets:**")
        for ex in report["examples"]:
            lines.append(
                f"  - `{ex['ticket_id']}` | {ex['subject']} ({ex['category']} / {ex['subcategory']} / {ex['subscription_tier']}) | tags={ex['tags']}"
            )
        lines.append("")
    return "\n".join(lines)


def fig_bar_top_scores(grouped: pd.DataFrame):
    top = grouped.head(10).copy()
    fig, ax = plt.subplots(figsize=(10, 5.5))
    ax.barh(top["opportunity_name"][::-1], top["opportunity_score"][::-1])
    ax.set_xlabel("Opportunity score")
    ax.set_ylabel("Opportunity")
    ax.set_title("Top opportunity scores")
    plt.tight_layout()
    return fig


def fig_heatmap(grouped: pd.DataFrame):
    top = grouped.head(8).copy()
    cols = [
        "volume_score",
        "severity_score",
        "premium_score",
        "virality_score",
        "product_dependency_score",
        "ops_burden_score",
        "backlog_score",
    ]
    labels = [
        "Volume",
        "Severity",
        "Premium",
        "Virality",
        "Prod dep.",
        "Ops burden",
        "Backlog",
    ]
    heat = top[["opportunity_name"] + cols].set_index("opportunity_name")
    fig, ax = plt.subplots(figsize=(11, 5.5))
    im = ax.imshow(heat.values, aspect="auto")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.set_yticks(range(len(heat.index)))
    ax.set_yticklabels(heat.index)
    ax.set_title("Priority heatmap by opportunity")
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    plt.tight_layout()
    return fig


def fig_tier_mix(df: pd.DataFrame, grouped: pd.DataFrame):
    top3_names = grouped.head(3)["opportunity_name"].tolist()
    subset = df[df["opportunity_name"].isin(top3_names)].copy()
    subset["subscription_tier_clean"] = subset["subscription_tier"].fillna("unknown").astype(str)
    pivot = pd.pivot_table(
        subset,
        index="opportunity_name",
        columns="subscription_tier_clean",
        values="ticket_id",
        aggfunc="count",
        fill_value=0,
    )
    fig, ax = plt.subplots(figsize=(11, 5.5))
    pivot.plot(kind="bar", ax=ax)
    ax.set_title("Subscription tier mix - top 3 opportunities")
    ax.set_xlabel("Opportunity")
    ax.set_ylabel("Ticket count")
    ax.tick_params(axis="x", rotation=20)
    plt.tight_layout()
    return fig


def fig_tag_distribution(tag_df: pd.DataFrame):
    fig, ax = plt.subplots(figsize=(10, 5.5))
    shown = tag_df.head(12).copy()
    ax.barh(shown["tag"][::-1], shown["count"][::-1])
    ax.set_xlabel("Count")
    ax.set_ylabel("Tag")
    ax.set_title("Most frequent tags in the dataset")
    plt.tight_layout()
    return fig


def analyze(df_raw: pd.DataFrame, rules: List[Dict], weights: Dict[str, float], include_tags: bool,
            premium_tiers: List[str], high_priority_values: List[str], open_status_values: List[str],
            virality_patterns: List[str], trust_patterns: List[str]):
    df = df_raw.copy()
    for col in EXPECTED_COLS:
        if col not in df.columns:
            df[col] = None

    df["first_response_time_min_num"] = df["first_response_time_min"].apply(lambda x: safe_float(x, 0))
    df["resolution_time_hours_num"] = df["resolution_time_hours"].apply(lambda x: safe_float(x, 0))
    df["csat_num"] = df["csat"].apply(lambda x: safe_float(x, np.nan))

    df["evidence_text"] = df.apply(lambda row: build_evidence_text(row, include_tags=include_tags), axis=1)
    classified = df["evidence_text"].apply(lambda txt: classify_opportunity(txt, rules)).apply(pd.Series)
    df = pd.concat([df, classified], axis=1)

    df["is_premium"] = df["subscription_tier"].apply(lambda x: premium_flag(x, premium_tiers))
    df["is_high_priority"] = df["priority"].apply(lambda x: high_priority_flag(x, high_priority_values))
    df["is_open"] = df["status"].apply(lambda x: open_status_flag(x, open_status_values))
    df["is_escalated_product"] = df["escalated_to_product"].apply(escalation_flag)
    df["virality_risk"] = df["evidence_text"].apply(lambda x: regex_hit_score(x, virality_patterns, 5))
    df["trust_risk"] = df["evidence_text"].apply(lambda x: regex_hit_score(x, trust_patterns, 5))
    df["severity_proxy"] = df.apply(severity_proxy, axis=1)
    df["tag_list"] = df["tags"].apply(parse_tags)
    df["has_tags"] = df["tag_list"].apply(lambda x: int(len(x) > 0))

    grouped = aggregate_opportunities(df, weights)
    reports = create_opportunity_reports(grouped, df)
    qa = diagnostics(df)
    tag_df = all_tag_counts(df)
    return df, grouped, reports, qa, tag_df


# -------------------------
# UI
# -------------------------
st.title("Finary Care Ops - Rule-based Opportunity Analyzer")
st.caption("Upload the Excel file, tune the rule-based weighting, and generate structured outputs for Product and Engineering.")

with st.sidebar:
    st.header("1) Upload")
    uploaded = st.file_uploader("Excel file (.xlsx)", type=["xlsx"])

    st.header("2) Scoring weights")
    volume_w = st.slider("Volume", 0.0, 1.0, 0.25, 0.01)
    severity_w = st.slider("Severity", 0.0, 1.0, 0.20, 0.01)
    premium_w = st.slider("Premium / revenue exposure", 0.0, 1.0, 0.15, 0.01)
    virality_w = st.slider("Virality / reputation risk", 0.0, 1.0, 0.15, 0.01)
    product_dep_w = st.slider("Product dependency / escalation", 0.0, 1.0, 0.10, 0.01)
    ops_burden_w = st.slider("Operational burden", 0.0, 1.0, 0.10, 0.01)
    backlog_w = st.slider("Backlog / open tickets", 0.0, 1.0, 0.05, 0.01)

    include_tags = st.checkbox("Use tags column in evidence text", value=True)

    st.header("3) Classification dictionaries")
    premium_tiers_text = st.text_area("Premium tiers", value="\n".join(PREMIUM_TIERS_DEFAULT), height=120)
    high_priority_text = st.text_area("High priority values", value="\n".join(HIGH_PRIORITY_VALUES_DEFAULT), height=100)
    open_status_text = st.text_area("Open status values", value="\n".join(OPEN_STATUS_VALUES_DEFAULT), height=120)

    st.header("4) Risk regex")
    virality_text = st.text_area("Virality patterns", value="\n".join(VIRALITY_PATTERNS_DEFAULT), height=220)
    trust_text = st.text_area("Trust risk patterns", value="\n".join(TRUST_RISK_PATTERNS_DEFAULT), height=180)

    st.header("5) Opportunity rules")
    st.caption("You can edit the regex list for each rule.")
    editable_rules = []
    for idx, rule in enumerate(DEFAULT_RULES):
        with st.expander(rule["name"], expanded=False):
            name = st.text_input(f"Rule name {idx+1}", value=rule["name"], key=f"name_{idx}")
            owner = st.text_input(f"Owner {idx+1}", value=rule["owner"], key=f"owner_{idx}")
            rtype = st.text_input(f"Type {idx+1}", value=rule["type"], key=f"type_{idx}")
            patterns_text = st.text_area(
                f"Patterns {idx+1} (one regex per line)",
                value="\n".join(rule["patterns"]),
                height=220,
                key=f"patterns_{idx}",
            )
            editable_rules.append(
                {
                    "name": name,
                    "owner": owner,
                    "type": rtype,
                    "patterns": parse_multiline_list(patterns_text),
                }
            )

weights = {
    "volume": volume_w,
    "severity": severity_w,
    "premium": premium_w,
    "virality": virality_w,
    "product_dependency": product_dep_w,
    "ops_burden": ops_burden_w,
    "backlog": backlog_w,
}

premium_tiers = parse_multiline_list(premium_tiers_text)
high_priority_values = parse_multiline_list(high_priority_text)
open_status_values = parse_multiline_list(open_status_text)
virality_patterns = parse_multiline_list(virality_text)
trust_patterns = parse_multiline_list(trust_text)

if uploaded is None:
    st.info("Upload the Excel file to start the analysis.")
    st.stop()

raw_df = read_excel_safe(uploaded.getvalue(), uploaded.name)

st.subheader("Dataset overview")
col1, col2, col3 = st.columns(3)
col1.metric("Rows loaded", len(raw_df))
col2.metric("Columns found", len(raw_df.columns))
col3.metric("Tags column present", "Yes" if "tags" in raw_df.columns else "No")

with st.expander("Show detected columns", expanded=False):
    st.write(list(raw_df.columns))

try:
    analyzed_df, grouped_df, reports, qa, tag_df = analyze(
        raw_df,
        editable_rules,
        weights,
        include_tags,
        premium_tiers,
        high_priority_values,
        open_status_values,
        virality_patterns,
        trust_patterns,
    )
except Exception as e:
    st.error(f"Analysis failed: {e}")
    st.stop()

st.subheader("Quality checks / evidence that all tickets were studied")
q1, q2, q3, q4, q5 = st.columns(5)
q1.metric("Total tickets", qa["total_tickets"])
q2.metric("Processed tickets", qa["processed_tickets"])
q3.metric("Coverage rate", f"{qa['coverage_rate']:.1%}")
q4.metric("Uncategorized", qa["uncategorized_tickets"])
q5.metric("Tag coverage", f"{qa['tag_coverage_rate']:.1%}")

q6, q7, q8, q9 = st.columns(4)
q6.metric("Matched rule rate", f"{qa['matched_rule_rate']:.1%}")
q7.metric("Duplicate ticket IDs", qa["duplicate_ticket_ids"])
q8.metric("Missing subject", qa["missing_subject"])
q9.metric("Missing description", qa["missing_description"])

if qa["coverage_rate"] < 1:
    st.warning("Some tickets were not processed. Review the input columns and parsing logic.")
if qa["uncategorized_rate"] > 0.20:
    st.warning("A high share of tickets is uncategorized. You may need to expand the rule set.")

st.subheader("Top opportunity table")
st.dataframe(grouped_df, use_container_width=True, hide_index=True)

chart_col1, chart_col2 = st.columns(2)
with chart_col1:
    st.pyplot(fig_bar_top_scores(grouped_df))
with chart_col2:
    st.pyplot(fig_heatmap(grouped_df))

chart_col3, chart_col4 = st.columns(2)
with chart_col3:
    st.pyplot(fig_tier_mix(analyzed_df, grouped_df))
with chart_col4:
    if not tag_df.empty:
        st.pyplot(fig_tag_distribution(tag_df))
    else:
        st.info("No tags found to display.")

st.subheader("Top tags from the last column")
st.dataframe(tag_df, use_container_width=True, hide_index=True)

st.subheader("Top 3 opportunities - Product / Engineering report")
for i, report in enumerate(reports, start=1):
    with st.container(border=True):
        st.markdown(f"### {i}. {report['opportunity_name']}")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Score", f"{report['score']:.2f}")
        c2.metric("Volume", report["ticket_volume"])
        c3.metric("Premium share", f"{report['premium_share']:.1%}")
        c4.metric("Escalation rate", f"{report['escalation_rate']:.1%}")
        c5.metric("Avg resolution", f"{report['avg_resolution_time_h']:.1f}h")

        st.code(render_report_text(report), language="text")

        if report["examples"]:
            st.markdown("**Example tickets**")
            ex_df = pd.DataFrame(report["examples"])
            st.dataframe(ex_df, use_container_width=True, hide_index=True)

st.subheader("Ticket-level audit trail")
audit_cols = [
    "ticket_id",
    "category",
    "subcategory",
    "subject",
    "status",
    "subscription_tier",
    "tags",
    "opportunity_name",
    "rule_strength",
    "matched_patterns",
    "all_candidate_opportunities",
    "severity_proxy",
    "virality_risk",
    "trust_risk",
]
st.dataframe(analyzed_df[audit_cols], use_container_width=True, hide_index=True)

markdown_report = build_markdown_reports(reports)

st.download_button(
    "Download opportunity scoring CSV",
    data=grouped_df.to_csv(index=False).encode("utf-8"),
    file_name="opportunity_scoring_table.csv",
    mime="text/csv",
)
st.download_button(
    "Download ticket-level audit CSV",
    data=analyzed_df.to_csv(index=False).encode("utf-8"),
    file_name="ticket_level_scored_output.csv",
    mime="text/csv",
)
st.download_button(
    "Download top 3 opportunity report (Markdown)",
    data=markdown_report.encode("utf-8"),
    file_name="top_3_opportunity_reports.md",
    mime="text/markdown",
)

st.caption("Tip: keep the QA section visible in your demo. It shows that every ticket was reviewed, how many were matched by rules, and how much evidence came from the tags column.")
